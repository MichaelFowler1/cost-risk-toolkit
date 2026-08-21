"""
app.py - The desktop tool: lot cost model with the added statistics.

The window an analyst sees is the one from the original tool -- the same four
tabs, the same fields, the same paste-from-Excel grids, producing the same
workbook. Nothing about the estimate has moved.

What is added sits alongside it. Tab 3 gains a block of settings for the extra
analytics, and a fifth tab reports what the point estimates on their own cannot
say: how much the log-space fit understates the mean, which analogy lot is
setting the slope, a prediction interval on every projected lot, and a risk
distribution on the total buy. All of it can be switched off, in which case the
tool behaves exactly as it did before.

Run it with ``python -m cost_core.gui`` or ``ce-core gui``.
"""

from __future__ import annotations

import os
import sys
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import numpy as np
import pandas as pd

from cost_core.lotmodel import (SETTINGS, enrich_run, generate_analyst_summary,
                                generate_fit_chart_data, run_lot_cost_model,
                                save_complete_excel_workbook)
from cost_core.lotmodel.enrich import DEFAULT_LOT_CORRELATION, EnrichmentError
from cost_core.gui.widgets import (EXAMPLE_ANALOGY, EXAMPLE_ESTIMATE, LotGrid,
                                   default_output_dir, parse_float, split_row)


class LotCostApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Lot Cost Model - Learning Curve / Rate Analysis")
        self.geometry("980x760")
        self.minsize(860, 640)

        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass
        style.configure("Head.TLabel", font=("Segoe UI", 9, "bold"))
        style.configure("Title.TLabel", font=("Segoe UI", 13, "bold"))
        style.configure("Sub.TLabel", foreground="#555")

        ttk.Label(
            self, text="Lot Cost Model", style="Title.TLabel"
        ).pack(anchor="w", padx=12, pady=(10, 0))
        ttk.Label(
            self,
            text=(
                "Enter historical analogy lots and forecast lots, then run the "
                "model. Paste directly from Excel with Ctrl+V."
            ),
            style="Sub.TLabel",
        ).pack(anchor="w", padx=12, pady=(0, 8))

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12)
        self.tab_analogy = ttk.Frame(nb)
        self.tab_estimate = ttk.Frame(nb)
        self.tab_run = ttk.Frame(nb)
        self.tab_results = ttk.Frame(nb)
        self.tab_stats = ttk.Frame(nb)
        nb.add(self.tab_analogy, text="  1. Analogy Lots  ")
        nb.add(self.tab_estimate, text="  2. Estimate Lots  ")
        nb.add(self.tab_run, text="  3. Run Info & Settings  ")
        nb.add(self.tab_results, text="  4. Results  ")
        nb.add(self.tab_stats, text="  5. Statistics  ")
        self.nb = nb

        self._build_analogy()
        self._build_estimate()
        self._build_runinfo()
        self._build_results()
        self._build_stats()
        self._build_actionbar()

    # -- tabs ---------------------------------------------------------------
    def _build_analogy(self):
        f = self.tab_analogy
        ttk.Label(
            f,
            text=(
                "Historical lots used to fit the curve. Need at least 3 lots "
                "with a unit cost.\n"
                "Leave AUC blank for a quantity-only lot (its units count "
                "toward learning, but it is not fit)."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=8, pady=(8, 6))

        self.grid_analogy = LotGrid(
            f,
            ["Fiscal Year", "Lot Quantity", "Unit Cost AUC ($K)"],
            [14, 14, 20],
        )
        self.grid_analogy.pack(fill="both", expand=True, padx=8)

        bar = ttk.Frame(f)
        bar.pack(fill="x", padx=8, pady=8)
        ttk.Button(
            bar, text="Add Row", command=self.grid_analogy.add_row
        ).pack(side="left")
        ttk.Button(
            bar, text="Delete Last Row", command=self.grid_analogy.delete_last
        ).pack(side="left", padx=4)
        ttk.Button(
            bar, text="Clear", command=self.grid_analogy.clear
        ).pack(side="left")
        ttk.Button(
            bar,
            text="Load Example",
            command=lambda: self.grid_analogy.load(EXAMPLE_ANALOGY),
        ).pack(side="right")

        for _ in range(6):
            self.grid_analogy.add_row()

    def _build_estimate(self):
        f = self.tab_estimate
        ttk.Label(
            f,
            text=(
                "Forecast lots to be costed. Complexity Factor is optional; a "
                "blank one carries the\nprevious lot's value forward (1.0 to "
                "start)."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=8, pady=(8, 6))

        self.grid_estimate = LotGrid(
            f,
            ["Fiscal Year", "Lot Quantity", "Complexity Factor"],
            [14, 14, 20],
        )
        self.grid_estimate.pack(fill="both", expand=True, padx=8)

        bar = ttk.Frame(f)
        bar.pack(fill="x", padx=8, pady=8)
        ttk.Button(
            bar, text="Add Row", command=self.grid_estimate.add_row
        ).pack(side="left")
        ttk.Button(
            bar,
            text="Delete Last Row",
            command=self.grid_estimate.delete_last,
        ).pack(side="left", padx=4)
        ttk.Button(
            bar, text="Clear", command=self.grid_estimate.clear
        ).pack(side="left")
        ttk.Button(
            bar,
            text="Load Example",
            command=lambda: self.grid_estimate.load(EXAMPLE_ESTIMATE),
        ).pack(side="right")

        for _ in range(8):
            self.grid_estimate.add_row()

    def _build_runinfo(self):
        f = self.tab_run
        box = ttk.LabelFrame(f, text="Run Info")
        box.pack(fill="x", padx=8, pady=10)

        self.var_runid = tk.StringVar(value=SETTINGS["DefaultRunID"])
        self.var_program = tk.StringVar(value=SETTINGS["DefaultProgram"])
        self.var_label = tk.StringVar(value=SETTINGS["DefaultRunLabel"])
        self.var_baseyear = tk.StringVar(value="")

        fields = [
            ("Run ID", self.var_runid, ""),
            ("Program", self.var_program, ""),
            ("Run label", self.var_label, ""),
            ("Base year ($)", self.var_baseyear, "blank = not stated"),
        ]
        for r, (lbl, var, hint) in enumerate(fields):
            ttk.Label(box, text=lbl + ":").grid(
                row=r, column=0, sticky="e", padx=8, pady=5
            )
            ttk.Entry(box, textvariable=var, width=42).grid(
                row=r, column=1, sticky="w", pady=5
            )
            if hint:
                ttk.Label(box, text=hint, style="Sub.TLabel").grid(
                    row=r, column=2, sticky="w", padx=8
                )

        box2 = ttk.LabelFrame(f, text="Model Settings")
        box2.pack(fill="x", padx=8, pady=6)

        self.var_costscale = tk.StringVar(
            value=str(SETTINGS["CostUnitScale"])
        )
        self.var_totalscale = tk.StringVar(value=str(SETTINGS["TotalScale"]))
        self.var_defaultcf = tk.StringVar(value=str(SETTINGS["DefaultCF"]))
        self.var_tgate = tk.StringVar(value=str(SETTINGS["TGate"]))
        self.var_fitprior = tk.StringVar(
            value=str(SETTINGS["FitPriorUnits"])
        )
        self.var_fcstprior = tk.StringVar(
            value=str(SETTINGS["FcstPriorUnits"])
        )
        self.var_legacyrate = tk.BooleanVar(
            value=SETTINGS["LegacyRateOmission"]
        )

        s_fields = [
            ("Cost unit scale", self.var_costscale, "1 = $K, 1000 = dollars"),
            ("Total scale", self.var_totalscale, "applied on top, for totals"),
            ("Default complexity", self.var_defaultcf, "used if none given"),
            ("t-gate", self.var_tgate, "significance cutoff for rate term"),
            ("Prior units (fit)", self.var_fitprior, "units built before lot 1"),
            ("Prior units (forecast)", self.var_fcstprior, ""),
        ]
        for r, (lbl, var, hint) in enumerate(s_fields):
            ttk.Label(box2, text=lbl + ":").grid(
                row=r, column=0, sticky="e", padx=8, pady=4
            )
            ttk.Entry(box2, textvariable=var, width=16).grid(
                row=r, column=1, sticky="w", pady=4
            )
            if hint:
                ttk.Label(box2, text=hint, style="Sub.TLabel").grid(
                    row=r, column=2, sticky="w", padx=8
                )
        ttk.Checkbutton(
            box2,
            text="Legacy: drop the rate term from Rate & LC+Rate "
                 "projections (matches the original tool, overstates cost)",
            variable=self.var_legacyrate,
        ).grid(row=len(s_fields), column=0, columnspan=3, sticky="w", padx=8, pady=6)

        # -- added statistics ------------------------------------------------
        # These never touch the estimate. They read the models the engine
        # already fitted and report how much confidence those numbers carry.
        box3 = ttk.LabelFrame(f, text="Added Statistics")
        box3.pack(fill="x", padx=8, pady=6)

        self.var_stats_on = tk.BooleanVar(value=True)
        self.var_level = tk.StringVar(value="0.80")
        self.var_iters = tk.StringVar(value="20000")
        self.var_seed = tk.StringVar(value="0")
        self.var_lotcorr = tk.StringVar(value=str(DEFAULT_LOT_CORRELATION))

        ttk.Checkbutton(
            box3,
            text="Compute added statistics (unbiased refits, intervals, "
                 "influence, buy risk)",
            variable=self.var_stats_on,
        ).grid(row=0, column=0, columnspan=3, sticky="w", padx=8, pady=(6, 2))

        x_fields = [
            ("Interval level", self.var_level, "0.80 = 80% prediction interval"),
            ("Risk iterations", self.var_iters, "Monte Carlo draws on the buy"),
            ("Seed", self.var_seed, "fixed, so the P80 is reproducible"),
            ("Lot correlation", self.var_lotcorr,
             "residual correlation across estimate lots"),
        ]
        for r, (lbl, var, hint) in enumerate(x_fields, start=1):
            ttk.Label(box3, text=lbl + ":").grid(
                row=r, column=0, sticky="e", padx=8, pady=4
            )
            ttk.Entry(box3, textvariable=var, width=16).grid(
                row=r, column=1, sticky="w", pady=4
            )
            ttk.Label(box3, text=hint, style="Sub.TLabel").grid(
                row=r, column=2, sticky="w", padx=8
            )

    def _build_results(self):
        f = self.tab_results
        self.lbl_result = ttk.Label(
            f,
            text="No run yet. Fill in the lots, then click Run Model.",
            style="Sub.TLabel",
        )
        self.lbl_result.pack(anchor="w", padx=8, pady=8)

        cols = ("Item", "Value", "LC", "Rate", "LC+Rate")
        self.tree = ttk.Treeview(f, columns=cols, show="headings", height=22)
        widths = (250, 330, 120, 120, 120)
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="w")
        vs = ttk.Scrollbar(f, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vs.set)
        vs.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self.tree.tag_configure("sel", background="#dff0d8")

    def _build_stats(self):
        f = self.tab_stats
        self.lbl_stats = ttk.Label(
            f,
            text=(
                "No run yet. These read the model the engine selected and "
                "report what the point estimates cannot say."
            ),
            style="Sub.TLabel",
            justify="left",
            wraplength=920,
        )
        self.lbl_stats.pack(anchor="w", padx=8, pady=8)

        cols = ("Section", "Item", "Value", "Note")
        self.tree_stats = ttk.Treeview(
            f, columns=cols, show="headings", height=22
        )
        for c, w in zip(cols, (150, 260, 150, 380)):
            self.tree_stats.heading(c, text=c)
            self.tree_stats.column(c, width=w, anchor="w")
        vs = ttk.Scrollbar(f, orient="vertical", command=self.tree_stats.yview)
        self.tree_stats.configure(yscrollcommand=vs.set)
        vs.pack(side="right", fill="y")
        self.tree_stats.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self.tree_stats.tag_configure("hdr", background="#e8eef5")
        self.tree_stats.tag_configure("warn", background="#fdf0e2")

    def _collect_stats_options(self) -> dict:
        """Read the added-statistics settings, refusing anything unusable."""
        def num(var, label, caster=float):
            try:
                return caster(var.get().strip())
            except ValueError:
                raise ValueError(f"Setting '{label}' must be a number.")

        level = num(self.var_level, "Interval level")
        if not 0.0 < level < 1.0:
            raise ValueError(
                f"Interval level must be between 0 and 1; got {level}. "
                "Use 0.80 for an 80% interval."
            )
        iters = num(self.var_iters, "Risk iterations", int)
        if iters < 2:
            raise ValueError("Risk iterations must be at least 2.")
        corr = num(self.var_lotcorr, "Lot correlation")
        if not -1.0 < corr < 1.0:
            raise ValueError(
                f"Lot correlation must be between -1 and 1; got {corr}."
            )
        return {
            "enabled": bool(self.var_stats_on.get()),
            "level": level,
            "n_iter": iters,
            "seed": num(self.var_seed, "Seed", int),
            "lot_correlation": corr,
        }

    def _show_stats(self, en, opts):
        """Fill the statistics tab from an Enrichment."""
        t = self.tree_stats
        t.delete(*t.get_children())

        def row(section, item, value, note="", tag=""):
            t.insert("", "end", values=(section, item, value, note),
                     tags=(tag,) if tag else ())

        m = en.methods
        row("Retransformation", "", "", "", "hdr")
        row("", "Selected model", en.selected_model,
            "Statistics below are for this model only.")
        row("", "OLS understates the mean by",
            f"{m.percent_understated:.3f}%",
            "Fitting ln(cost) and exponentiating back estimates the median, "
            "not the mean.")
        row("", "Theoretical factor exp(s2/2)", f"{m.theoretical_factor:.5f}",
            "Under lognormal log-space errors.")
        row("", "Duan smearing factor", f"{m.smearing_factor:.5f}",
            "Nonparametric. Agreement with the line above means the lognormal "
            "assumption is doing no work.")
        for _, r in m.frame.iterrows():
            b = "" if pd.isna(r["b (learning)"]) else f"{r['b (learning)']:.6f}"
            c = "" if pd.isna(r["c (rate)"]) else f"{r['c (rate)']:.6f}"
            row("", f"{r['Method']} fit",
                f"T1 {r['T1 ($K)']:,.2f}",
                f"b {b}  c {c}   mean % error {r['Mean % error']:+.2e}   "
                f"MAPE {r['MAPE']:.2%}")

        row("Influence", "", "", "", "hdr")
        row("", "Leverage flag", f"> {2 * 2 / max(len(en.influence), 1):.3f}",
            "Conventional 2p/n. A flag, not a verdict.")
        for _, r in en.influence.iterrows():
            flags = []
            if r["High leverage"]:
                flags.append("high leverage")
            if r["Influential"]:
                flags.append("INFLUENTIAL")
            # Pulled out of the f-string: a backslash inside an f-string
            # expression is a syntax error before Python 3.12, and this
            # package supports 3.11.
            leverage = float(r["Leverage"])
            cooks = float(r["Cook's D"])
            note = f"leverage {leverage:.3f}   Cook's D {cooks:.3f}"
            if flags:
                note += "   " + ", ".join(flags)
            row("", str(r["Lot"]), f"{r['% error']:+.2f}%", note,
                "warn" if flags else "")

        row("Prediction intervals", "", "", "", "hdr")
        row("", "Level", f"{opts['level']:.0%}",
            "A new lot, not the mean of the fitted line. Carries the residual "
            "scatter, which more analogy lots will not remove.")
        for _, r in en.intervals.iterrows():
            row("", f"Lot {int(r['Lot'])} (FY {r['Fiscal Year']:.0f})"
                if pd.notna(r["Fiscal Year"]) else f"Lot {int(r['Lot'])}",
                f"${r['Lot Cost ($)']:,.0f}",
                f"${r['Lot Cost Lower']:,.0f} to ${r['Lot Cost Upper']:,.0f}")

        risk = en.risk
        row("Buy risk", "", "", "", "hdr")
        row("", "Point estimate", f"${risk.point_estimate:,.0f}",
            f"Sits at the {risk.point_estimate_percentile:.0f}th percentile "
            f"of the risk distribution.")
        row("", "P50", f"${risk.p50:,.0f}", "")
        row("", "P80", f"${risk.p80:,.0f}",
            f"Reserve of ${risk.p80 - risk.point_estimate:,.0f} "
            f"({100 * (risk.p80 / risk.point_estimate - 1):.1f}%)")
        row("", "P90", f"${risk.p90:,.0f}", "")
        row("", "CV of the total", f"{risk.cv:.2%}",
            f"{risk.n_iter:,} draws, t with {risk.dof} degrees of freedom, "
            f"lot residuals correlated at {risk.lot_correlation:.2f}.")

        if en.warnings_raised:
            row("Cautions", "", "", "", "hdr")
            for w in en.warnings_raised:
                row("", "", "", w, "warn")

        self.lbl_stats.config(
            text=(
                f"{en.selected_model} model. " + risk.narrative()
            )
        )

    def _build_actionbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=12, pady=10)

        ttk.Label(bar, text="Save workbook to:").pack(side="left")
        self.var_outfile = tk.StringVar(
            value=os.path.join(
                default_output_dir(), "Lot_Cost_Model_Complete_Suite.xlsx"
            )
        )
        ttk.Entry(bar, textvariable=self.var_outfile).pack(
            side="left", fill="x", expand=True, padx=6
        )
        ttk.Button(bar, text="Browse...", command=self._browse).pack(
            side="left"
        )
        self.btn_run = ttk.Button(
            bar, text="Run Model", command=self.run_model
        )
        self.btn_run.pack(side="left", padx=(10, 0))

        self.var_status = tk.StringVar(value="Ready.")
        ttk.Label(
            self, textvariable=self.var_status, style="Sub.TLabel"
        ).pack(anchor="w", padx=12, pady=(0, 8))

    def _browse(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=os.path.basename(self.var_outfile.get()),
            initialdir=os.path.dirname(self.var_outfile.get())
            or default_output_dir(),
            title="Save model workbook as",
        )
        if path:
            self.var_outfile.set(path)

    # -- input parsing ------------------------------------------------------
    def _collect_analogy(self) -> pd.DataFrame:
        rows = self.grid_analogy.get_rows()
        if not rows:
            raise ValueError("No analogy lots entered (tab 1).")

        fy, qty, auc = [], [], []
        for i, r in enumerate(rows, start=1):
            try:
                q = parse_float(r[1])
            except ValueError:
                raise ValueError(
                    f"Analogy row {i}: Lot Quantity '{r[1]}' is not a number."
                )
            if q <= 0:
                raise ValueError(
                    f"Analogy row {i}: Lot Quantity must be greater than 0."
                )
            try:
                y = parse_float(r[0]) if r[0] else np.nan
            except ValueError:
                raise ValueError(
                    f"Analogy row {i}: Fiscal Year '{r[0]}' is not a number."
                )
            if r[2]:
                try:
                    c = parse_float(r[2])
                except ValueError:
                    raise ValueError(
                        f"Analogy row {i}: Unit Cost '{r[2]}' is not a number."
                    )
                if c <= 0:
                    raise ValueError(
                        f"Analogy row {i}: Unit Cost must be greater than 0 "
                        "(leave it blank for a quantity-only lot)."
                    )
            else:
                c = np.nan
            fy.append(y)
            qty.append(q)
            auc.append(c)

        n_costed = sum(1 for c in auc if pd.notna(c))
        if n_costed < 3:
            raise ValueError(
                "The learning curve needs at least 3 analogy lots with both a "
                f"quantity and a unit cost. Found {n_costed}."
            )

        return pd.DataFrame(
            {
                "Lot": list(range(1, len(rows) + 1)),
                "Lot FY": fy,
                "Qty": qty,
                "AUC ($K)": auc,
            }
        )

    def _collect_estimate(self) -> pd.DataFrame:
        rows = self.grid_estimate.get_rows()
        if not rows:
            raise ValueError("No estimate lots entered (tab 2).")

        fy, qty, cf = [], [], []
        for i, r in enumerate(rows, start=1):
            try:
                q = parse_float(r[1])
            except ValueError:
                raise ValueError(
                    f"Estimate row {i}: Lot Quantity '{r[1]}' is not a number."
                )
            if q <= 0:
                raise ValueError(
                    f"Estimate row {i}: Lot Quantity must be greater than 0."
                )
            try:
                y = parse_float(r[0]) if r[0] else np.nan
            except ValueError:
                raise ValueError(
                    f"Estimate row {i}: Fiscal Year '{r[0]}' is not a number."
                )
            if r[2]:
                try:
                    c = parse_float(r[2])
                except ValueError:
                    raise ValueError(
                        f"Estimate row {i}: Complexity Factor '{r[2]}' is not "
                        "a number."
                    )
            else:
                c = np.nan
            fy.append(y)
            qty.append(q)
            cf.append(c)

        return pd.DataFrame(
            {
                "Lot": list(range(1, len(rows) + 1)),
                "Lot FY": fy,
                "Qty": qty,
                "Complexity": cf,
            }
        )

    def _collect_overrides(self) -> dict:
        def num(var, label, caster=float):
            try:
                return caster(var.get().strip())
            except ValueError:
                raise ValueError(f"Setting '{label}' must be a number.")

        return {
            "CostUnitScale": num(self.var_costscale, "Cost unit scale"),
            "TotalScale": num(self.var_totalscale, "Total scale"),
            "DefaultCF": num(self.var_defaultcf, "Default complexity"),
            "TGate": num(self.var_tgate, "t-gate"),
            "FitPriorUnits": num(self.var_fitprior, "Prior units (fit)", int),
            "FcstPriorUnits": num(
                self.var_fcstprior, "Prior units (forecast)", int
            ),
            "LegacyRateOmission": bool(self.var_legacyrate.get()),
        }

    def _save_workbook(self, path, proj, summ, chart, extras=None) -> str | None:
        """Save, re-prompting if the path is locked or not writable."""
        while True:
            try:
                save_complete_excel_workbook(path, proj, summ, chart)
                if extras:
                    self._append_extra_sheets(path, extras)
                return path
            except PermissionError:
                retry = messagebox.askretrycancel(
                    "Cannot write the file",
                    f"Could not write to:\n{path}\n\n"
                    "The file may be open in Excel, or the folder may be "
                    "read-only.\n\nClose the file and click Retry, or Cancel "
                    "to choose a different location.",
                )
                if retry:
                    continue
                new = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel workbook", "*.xlsx")],
                    initialfile=os.path.basename(path),
                    initialdir=default_output_dir(),
                    title="Save model workbook as",
                )
                if not new:
                    return None
                path = new
                self.var_outfile.set(path)

    # -- run ----------------------------------------------------------------
    def run_model(self):
        self.btn_run.config(state="disabled")
        self.var_status.set("Running...")
        self.update_idletasks()
        try:
            analogy_df = self._collect_analogy()
            estimate_df = self._collect_estimate()
            overrides = self._collect_overrides()

            run_info = {
                "RunID": self.var_runid.get().strip()
                or SETTINGS["DefaultRunID"],
                "Program": self.var_program.get().strip()
                or SETTINGS["DefaultProgram"],
                "RunLabel": self.var_label.get().strip()
                or SETTINGS["DefaultRunLabel"],
                "BaseYear": self.var_baseyear.get().strip(),
            }

            stats_opts = self._collect_stats_options()

            projections_df, models_ctx = run_lot_cost_model(
                analogy_df, estimate_df, overrides
            )
            summary_df = generate_analyst_summary(models_ctx, run_info)
            chart_df = generate_fit_chart_data(models_ctx)

            # The added statistics read the fitted models; they never feed
            # back into them, so a failure here cannot change the estimate.
            enrichment, extra_sheets = None, None
            if stats_opts["enabled"]:
                try:
                    enrichment = enrich_run(
                        models_ctx, projections_df, summary_df,
                        level=stats_opts["level"],
                        n_iter=stats_opts["n_iter"],
                        seed=stats_opts["seed"],
                        lot_correlation=stats_opts["lot_correlation"],
                    )
                    extra_sheets = enrichment.sheets()
                except EnrichmentError as exc:
                    messagebox.showwarning(
                        "Statistics unavailable",
                        f"The estimate ran and will be saved. The added "
                        "statistics could not be computed:"
                        + chr(10) + chr(10) + str(exc),
                    )

            path = self.var_outfile.get().strip()
            if not path:
                raise ValueError("Choose an output file first.")
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            saved = self._save_workbook(
                path, projections_df, summary_df, chart_df, extra_sheets
            )

            self._show_results(summary_df)
            if enrichment is not None:
                self._show_stats(enrichment, stats_opts)
            else:
                self.tree_stats.delete(*self.tree_stats.get_children())
                self.lbl_stats.config(
                    text="Added statistics were switched off for this run."
                )
            if saved:
                self.var_outfile.set(saved)
                self.var_status.set(f"Saved: {saved}")
                self.lbl_result.config(
                    text=(
                        f"Run complete. Workbook saved to:\n{saved}\n"
                        "Sheets: Analyst_Summary, Estimate_Projections, "
                        "Fit_Chart_Data (with 3 embedded charts)."
                    )
                )
                if messagebox.askyesno(
                    "Run complete",
                    f"Model ran successfully.\n\nSaved to:\n{saved}\n\n"
                    "Open the workbook now?",
                ):
                    try:
                        os.startfile(saved)
                    except Exception:
                        pass
            else:
                self.var_status.set("Run complete, workbook not saved.")
                self.lbl_result.config(
                    text="Run complete, but the workbook was not saved."
                )

        except ValueError as exc:
            messagebox.showerror("Check your input", str(exc))
            self.var_status.set("Input error.")
        except Exception as exc:
            messagebox.showerror(
                "Model error",
                f"{type(exc).__name__}: {exc}\n\n"
                f"{traceback.format_exc(limit=3)}",
            )
            self.var_status.set("Run failed.")
        finally:
            self.btn_run.config(state="normal")

    def _append_extra_sheets(self, path, extras: dict):
        """Add the statistics sheets to the workbook the engine just wrote.

        Appended rather than written in one pass so that the three original
        sheets are byte-for-byte what the old tool produced, and an analyst who
        wants only those can ignore the rest.
        """
        import openpyxl

        wb = openpyxl.load_workbook(path)
        for name, frame in extras.items():
            if name in wb.sheetnames:
                del wb[name]
            ws = wb.create_sheet(title=name[:31])
            ws.append([str(c) for c in frame.columns])
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
            for _, r in frame.iterrows():
                ws.append([
                    v.item() if hasattr(v, "item") else v
                    for v in r.tolist()
                ])
            for i, col in enumerate(frame.columns, start=1):
                width = max(len(str(col)) + 2, 14)
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(i)
                ].width = min(width, 40)
        wb.save(path)

    def _show_results(self, summary_df: pd.DataFrame):
        self.tree.delete(*self.tree.get_children())
        for _, row in summary_df.iterrows():
            vals = [
                str(row["Item"]),
                str(row["Value"]),
                str(row["LC"]),
                str(row["Rate"]),
                str(row["LC+Rate"]),
            ]
            tag = "sel" if vals[0] == "SELECTED" else ""
            self.tree.insert("", "end", values=vals, tags=(tag,))
        self.nb.select(self.tab_results)


def main(argv: list[str] | None = None) -> int:
    """Launch the desktop tool.

    Returns a non-zero exit code with a readable message when there is no
    desktop session, rather than a bare Tcl traceback.
    """
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)   # crisper text on high-DPI
    except Exception:
        pass

    try:
        app = LotCostApp()
    except tk.TclError as exc:
        print(f"Could not start the GUI: {exc}", file=sys.stderr)
        print("A desktop session is required to run this tool.", file=sys.stderr)
        return 1

    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
